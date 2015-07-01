__author__ = 'Quan'

import requests as r
from bs4 import BeautifulSoup as bs
import openpyxl as x

fn ='D:\LinhTinh\Code\PythonProject\sample.xlsx'
wb = x.Workbook()
wb = x.load_workbook(fn)

def getcafef():
cf = r.get('http://cafef.vn/')
soup1 = bs(cf.content)
	news = soup1.find_all("div", {"class":"right"})
	news = news[1]
	lis = news.find_all("li")
	li_content = [ele.text.strip() for ele in lis]
	ws=wb.worksheets[0]
for i in range(len(li_content)):
    ws.cell(row = i + 2, column = 2).value = li_content[i]
wb.save(fn)

def getvietstock():
	vs = wb.worksheets[0]
	vs=get('http://vietstock.vn')
	soup1=bs(vs.content)
	news=soup1.find_all("ul",{"class":"ui-tabs-nav"})
	lis=news.find_all("li")
	li_content=[ele.text.strip() for ele in lis]
	ws=wb.worksheets[0]
for i in range(len(li_content)):
    ws.cell(row = i + 10, column = 2).value = li_content[i]
wb.save(fn)


getcafef()
getvietstock()
	
	
	