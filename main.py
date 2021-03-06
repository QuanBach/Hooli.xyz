

import requests as r
from bs4 import BeautifulSoup as bs
import openpyxl as x


fn='C:\Users\Quan\Desktop\Book1.xlsx' #file name de gan duong dan vao
wb = x.Workbook()
wb = x.load_workbook(fn)
# def getFileName(): #lay ten file de xac dinh file nao` can sua
#     global fn
#     global wb
#     fn = tkfd.askopenfilename(filetypes=[('xlsx files', '.xlsx'), ('all files', '.*')], defaultextension='.xlsx')
#     txtBox.delete(0, tk.END)
#     txtBox.insert(0, fn)
#     wb = x.load_workbook(fn)

def getBasicPrice(): #tao ra 1 cai function nho
#    global wb #trong
#    global fn


    ws = wb.worksheets[0]

    vs = r.get('http://vietstock.vn/')

    soup1 = bs(vs.content)
    vnindex = soup1.find("div",{"id" : "PRICE_VNIndex","class":"header_middle_text2"})
    vnindex = vnindex.text
    ws['C5'] = vnindex

    hnxindex = soup1.find("div",{"id": "PRICE_HastcIndex"})
    hnxindex = hnxindex.text
    ws['C6'] = hnxindex

    oil = soup1.find("div",{"id": "PRICE_CL"})
    oil=oil.text
    ws['C7'] = oil

    wb.save(fn)


def getExchangeRate():
    global wb
    global fn

    # cafef_html = html.fromstring(cafef.content)
    # er = cafef_html.get_element_by_id('div_tygia')
    # print(type(er.text_content()))
    ws = wb.worksheets[0]

    cafef = r.get('http://cafef.vn/')

    soup1 = bs(cafef.content)
    div = soup1.find('div', {'id':'div_tygia'})
    table_body = div.find('tbody')
    rows = table_body.find_all('tr')


    for i in range(len(rows)):
        cols = rows[i].find_all('td')
        cols = [ele.text.strip() for ele in cols]
        for j in range(len(cols)):
            ws.cell(row=10+i, column=2+j).value = cols[j]

    wb.save(fn)
def getFinanceReport():
    try:
        fr = r.get('http://s.cafef.vn/bao-cao-tai-chinh/DPM/IncSta/2050/1/0/0/ket-qua-hoat-dong-kinh-doanh-tong-cong-ty-phan-bon-va-hoa-chat-dau-khictcp.chn')
    except r.ConnectionError:
        return False
    soup = bs(fr.content)
    table_header = soup.find('table', {'id':'tblGridData'})
    rows = table_header.find_all('tr')
    ws = wb.worksheets[1]
    for i in range(len(rows)):
        cols = rows[i].find_all('td')
        cols = [ele.text.strip() for ele in cols]
        for j in range(len(cols)):
            ws.cell(row=10+i, column=2+j).value = cols[j]

    table_content = soup.find('table', {'id':'tableContent'})
    rows = table_content.find_all('tr')
    rowid = 0
    for i in range(len(rows)):
        cols = rows[i].find_all('td')
        cols = [ele.text.strip() for ele in cols]
        if i % 3 == 0:
            for j in range(len(cols)):
                cols[j] = cols[j].replace(',', '')
                try:
                    cols[j]=int(cols[j])
                except ValueError:
                    pass
                ws.cell(row=11+rowid, column=2+j).value=cols[j]
            rowid+=1
    wb.save(fn)
    return True

def getNews():
    try:
        newsPage = r.get('http://s.cafef.vn/hose/DPM-tong-cong-ty-phan-bon-va-hoa-chat-dau-khictcp.chn')
    except r.ConnectionError:
        return False

    soup = bs(newsPage.content)
    div = soup.find('div', {'id':'divTopEvents'})
    lis = div.find_all('li')
    alis = div.find_all('a',href=True)
    ws = wb.worksheets[1]
    li_content = [ele.text.strip() for ele in lis] # lay text trong moi cai li
    for i in range(len(li_content)):
        ws.cell(row = i + 1, column = 1).value = li_content[i]
        ws.cell(row = i + 1, column = 3).value = 'http://s.cafef.vn'+alis[i]['href']
    wb.save(fn)
    return True

# def update():
#     global fn
#     if fn:
#         if not getBasicPrice():
#             stringAnnouncement.set('Cannot connect! Check your connection!')
#             return
#         if not getExchangeRate():
#             stringAnnouncement.set('Cannot connect! Check your connection!')
#             return
#         if not getFinanceReport():
#             stringAnnouncement.set('Cannot connect! Check your connection!')
#             return
#         if not getNews():
#             stringAnnouncement.set('Cannot connect! Check your connection!')
#             return
#         stringAnnouncement.set('Update successfully!')


# root = tk.Tk()
# stringFile = tk.StringVar()
# stringFile.set('File: ')
# stringAnnouncement = tk.StringVar()
# openBtn = tk.Button(root, text='...', command=getFileName)
# updateBtn = tk.Button(root, text='Update', command=update)
# fileLabel = tk.Label(root, textvariable=stringFile)
# txtBox = tk.Entry(root, width=50)
# announcementLabel = tk.Label(root, textvariable=stringAnnouncement)
#
# fileLabel.grid(row=0, column=0, padx=5, pady=5)
# openBtn.grid(row=0, column=2, padx=5, pady=5)
# txtBox.grid(row=0, column=1, padx=5, pady=5)
# updateBtn.grid(row=1, column=1, padx=5, pady=5)
# announcementLabel.grid(row=2, column=0, columnspan=3)
#
# root.mainloop()

getBasicPrice()
getExchangeRate()
getFinanceReport()
getNews()