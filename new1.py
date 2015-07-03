__author__ = 'Quan'

from bs4 import BeautifulSoup as bs
import requests as r
import openpyxl as x

fn='sample.xlsx'
wb= x.Workbook()
wb= x.load_workbook(fn)


def getcafef():
    cafef= r.get('http://cafef.vn/')
    cafef=bs(cafef.content)

    cafefnews=cafef.find_all("div",{"class": "right"})
    cafefnews=cafefnews[1] # chon div class right thu 2 trong tat cac cach div class right
    lis=cafefnews.find_all("li")
    linklis=cafefnews.find_all('a', href=True)
    li_content=[ele.text.strip() for ele in lis]

    cafefheadnews=cafef.find_all("div",{"class":"left"})
    cafefheadnews=cafefheadnews[1]
    linkcafefhead=cafefheadnews.find('a', href=True)
    cafefheadnews=cafefheadnews.find("h2").text

    ws=wb.worksheets[0]
    for i in range(len(li_content)):
        ws.cell(row = i + 3, column = 1).value = li_content[i]
        ws.cell(row = i + 3, column = 2).value = 'http://cafef.vn'+linklis[i]['href']
    ws.cell(row=2, column=1).value =cafefheadnews
    ws.cell(row=2, column=2).value ='http://cafef.vn'+linkcafefhead['href']
    wb.save(fn)



def getvietstock():
    vietstock= r.get('http://vietstock.vn/')
    vietstock=bs(vietstock.content)

    vietstocknews=vietstock.find_all("div",{"id": "hotnews_news"})
    vietstocknews=vietstocknews[0] # chon div class right thu 1 trong tat cac cach div class right
    lis=vietstocknews.find_all("li")
    li_content=[ele.text.strip() for ele in lis]
    linkvietnews=vietstocknews.find_all('a', href=True)

    vietstockhead=vietstock.find("div",{"class":"hotnews_head"})
    linkvietstockhead=vietstockhead.find('a', href=True)
    vietstockhead=vietstockhead.find("h1").text

    ws=wb.worksheets[0]
    for i in range(len(li_content)):
        ws.cell(row = i + 16, column = 1).value = li_content[i]
        ws.cell(row = i +16, column = 2).value = 'http://vietstock.vn'+linkvietnews[i]['href']
    ws.cell(row= 15, column=1).value = vietstockhead
    ws.cell(row=15, column=2).value = 'http://vietstock.vn'+linkvietstockhead['href']
    wb.save(fn)

def getgafin():
    nhipcaudautu=r.get('http://nhipcaudautu.vn/')
    nhipcaudautu=bs(nhipcaudautu.content)

    nhipcaunews=nhipcaudautu.find("ul",{"class":"homeUp"})
    listnews=nhipcaunews.find_all("ul")
    listnews=listnews[1]
    linklistnews=listnews.find_all('a',href=True)
    listnews=listnews.find_all("li")
    li_content2=[ele.text.strip() for ele in listnews]

    nhipcaunewshead=nhipcaunews.find("h1")
    textcuatitle=nhipcaunewshead.text
    linkdautien=nhipcaunewshead.find("a",href=True)

    ws=wb.worksheets[0]
    for i in range(len(li_content2)):
        ws.cell(row = i + 27, column =1).value = li_content2[i]
        ws.cell(row = i + 27, column = 2).value= 'http://nhipcaudautu.vn'+linklistnews[i]['href']
    ws.cell(row=26, column=1).value =textcuatitle
    ws.cell(row=26,column=2).value='http://nhipcaudautu.vn'+linkdautien['href']
    wb.save(fn)

def getADB():#41
    ADB= r.get('http://asianbondsonline.adb.org/vietnam/news.php')
    ADB=bs(ADB.content)

    ADBnews=ADB.find('table')
    ADBlink=ADBnews.find_all('a',href=True)
    ADBnews=ADBnews.find_all("tr")

    ws = wb.worksheets[0]
    for i in range(len(ADBnews)):
        cols = ADBnews[i].find_all('td',{"class":'title'})
        cols = [ele.text.strip() for ele in cols]

        ws.cell(row=41+i, column=2).value = ADBlink[i]['href']
        for j in range(len(cols)):

            ws.cell(row=41+i, column=1).value = cols[j]

    wb.save(fn)






getcafef()
getvietstock()
getgafin()
getADB()
